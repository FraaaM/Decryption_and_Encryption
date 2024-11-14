import hashlib
import os
import random
import string
import subprocess

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

chosenFileForEncryption = None

def chooseFileForDecryption():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        file_name = os.path.basename(file_path)
        file_label.set(f"Выбранный для расшифровки файл: {file_name}")
        btn_deobfuscate.config(state=tk.NORMAL)
    else:
        file_label.set("Файл для расшифровки не выбран!")
        messagebox.showinfo("Предупреждение","Файл для расшифровки не выбран!")

def chooseFileForEncryption():
    global chosenFileForEncryption
    chosenFileForEncryption = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    if chosenFileForEncryption:
        file_name = os.path.basename(chosenFileForEncryption)
        encrypt_file_label.set(f"Выбранный для шифрования файл: {file_name}")
        
        for button in (sha1_button, sha256_button, sha512_button, md5_button):
            button.config(state=tk.NORMAL)
    else:
        encrypt_file_label.set("Файл для шифрования не выбран!")
        messagebox.showinfo("Предупреждение", "Файл для шифрования не выбран!")

def retrieveHashData(file_path):
    df = pd.read_excel(file_path)
    hashes = df.iloc[:, 0].dropna().values
    
    base_directory = os.path.dirname(os.path.abspath(__file__))
    hashes_file_path = os.path.join(base_directory, "hashcat-6.2.6", "hashes_for_hashcat.txt")

    with open(hashes_file_path, "w") as hash_file:
        for hash_value in hashes:
            hash_file.write(f"{hash_value}\n")
    
    return len(hashes), hashes_file_path

def executeHashcat(file_path):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(base_dir, "output.txt")
    hashcat_executable = os.path.join(base_dir, "hashcat-6.2.6", "hashcat.exe")

    _, hashes_file = retrieveHashData(file_path)

    hashcat_command = [
        hashcat_executable, "-a", "3", "-m", "0", "-o", output_file, hashes_file, "?d" * 11 # маска
    ]
    
    process_result = subprocess.run(hashcat_command, cwd=os.path.join(base_dir, "hashcat-6.2.6"), capture_output=True, text=True)
    print(process_result.stdout)

    return output_file

def findSalts(known_numbers, decrypted_numbers):
        possible_salts = set()
        for decrypted in decrypted_numbers.values():
            try:
                poss_salt = int(decrypted) - known_numbers[0]
                if poss_salt < 0:
                    continue
                
                if all(str(num + poss_salt) in decrypted_numbers.values() for num in known_numbers):
                    possible_salts.add(poss_salt)
                    
            except ValueError:
                continue
        return list(possible_salts)

def applySaltForDecryption(file_path, decryptedfile_path):
    output_phones_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "decoded_numbers.xlsx")

    df = pd.read_excel(file_path)
    known_numbers = df.iloc[:, 2].dropna().astype(str).tolist()
    print(f"Debug data`s type from excel: {known_numbers}")
    known_numbers = [int(float(num)) if '.' in num else int(num) for num in known_numbers]
    print(f"Known numbers: {known_numbers}")

    decrypted_numbers = {}
    with open(decryptedfile_path, 'r') as file:
        for line in file:
            hash_value, decrypted_num = line.strip().split(':')
            decrypted_numbers[hash_value] = decrypted_num

    salts = findSalts(known_numbers, decrypted_numbers)
    if not salts:
        messagebox.showerror("Ошибка", "Соль не найдена")
        return

    if len(salts) == 1:
        selected_salt = salts[0]
        decrypted_list = []

        for hash_value, decrypted in decrypted_numbers.items():
            final_number = str(int(decrypted) - selected_salt)
            decrypted_list.append({"Хеши": hash_value, "Расшифрованные номера": final_number})

        result_df = pd.DataFrame(decrypted_list)
        result_df.to_excel(output_phones_path, index=False)

        workbook = load_workbook(output_phones_path)
        worksheet = workbook.active

        worksheet['C1'] = "Соль"
        worksheet['C1'].font = Font(bold=True)
        worksheet['C1'].alignment = Alignment(horizontal="center")
        worksheet['C2'] = selected_salt

        worksheet.column_dimensions['B'].width = len("Расшифрованные номера") + 5
        for col_cells in worksheet.columns:
            if col_cells[0].column_letter != 'B':
                max_length = max(len(str(cell.value)) for cell in col_cells if cell.value)
                worksheet.column_dimensions[col_cells[0].column_letter].width = max_length + 4

        workbook.save(output_phones_path)
        messagebox.showinfo("Вывод", f"Результат сохранен в decoded_numbers.xlsx")
    else:
        with pd.ExcelWriter(output_phones_path, engine='openpyxl', mode='w') as writer:
            salts_df = pd.DataFrame({"Соли": salts})
            salts_df.to_excel(writer, index=False)

        messagebox.showinfo("Вывод", f"Найдено несколько солей! Расшифровка не проводилась. Значения сохранены в decoded_numbers.xlsx")

def initiateDecryptionProcess():
    selected_file_path = file_label.get().replace("Выбранный для расшифровки файл: ", "")
    decrypted_output_path = executeHashcat(selected_file_path)  
    applySaltForDecryption(selected_file_path, decrypted_output_path)  


def generateSalt(salt_length):
    salt_characters = string.ascii_letters + string.digits
    return ''.join(random.choices(salt_characters, k=salt_length))

def encryptPhoneNumbers(hash_function, salt_length, salt_type='random'):
    data_frame = pd.read_excel(chosenFileForEncryption)
    phone_list = data_frame.iloc[:, 1].dropna().astype(str).tolist()
    
    #salt_length = 2
    salt_value = generateSalt(salt_length) if salt_type == 'random' else salt_type
    hashcat_pattern = f"{'?d' * 11}{'?a' * salt_length}"
    
    hash_file_path = os.path.join('hashcat-6.2.6', 'hashes.txt')
    os.makedirs(os.path.dirname(hash_file_path), exist_ok=True)

    with open(hash_file_path, 'w') as hash_file:
        hashed_entries = [
            {"Хеш": hash_function((phone_number + salt_value).encode()).hexdigest(), "Исходный номер": phone_number}
            for phone_number in phone_list
        ]
        hash_file.writelines(f"{entry['Хеш']}\n" for entry in hashed_entries)

    # подготовка данных для сохранения в Excel
    algorithm_title = hash_function.__name__.lower().replace('openssl_', '')
    hashcat_mode_code = {'md5': 0, 'sha1': 100, 'sha256': 1400, 'sha512': 1700}.get(algorithm_title, 'unknown')
    hashcat_command_str = f"hashcat -m {hashcat_mode_code} -a 3 -o found.txt hashes.txt {hashcat_pattern}"
    
    hashed_data_frame = pd.DataFrame(hashed_entries)
    hashed_data_frame.at[0, 'Соль'] = salt_value
    hashed_data_frame.at[0, 'Код hashcat'] = hashcat_command_str

    output_filename = f"numbers_{algorithm_title}.xlsx"
    output_filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), output_filename)
    hashed_data_frame.to_excel(output_filepath, index=False)

    workbook_instance = load_workbook(output_filepath)
    worksheet_instance = workbook_instance.active
    for col in worksheet_instance.columns:
        max_width = max(len(str(cell.value)) for cell in col if cell.value) + 2
        worksheet_instance.column_dimensions[col[0].column_letter].width = max_width

    workbook_instance.save(output_filepath)
    messagebox.showinfo("Вывод", f"Зашифрованные мобильные номера сохранены в {output_filename}")

salt_length = 1

def hash_SHA1():
    encryptPhoneNumbers(hashlib.sha1, salt_length)

def hash_SHA256():
    encryptPhoneNumbers(hashlib.sha256, salt_length)

def hash_SHA512():
    encryptPhoneNumbers(hashlib.sha512, salt_length)

def hash_MD5():
    encryptPhoneNumbers(hashlib.md5, salt_length) 


def setWindow(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

root = tk.Tk()
root.title("Телефонные номера: Шифрование и Расшифровка")
root.configure(bg="#F3F4F6")

# Стили
style = ttk.Style()
style.configure('TFrame', background="#F3F4F6")
style.configure('TLabel', background="#F3F4F6", font=('Arial', 11))
style.configure('TButton', font=('Arial', 11), padding=5)

frame_files = ttk.Frame(root, padding=10)
frame_files.grid(row=0, column=0, sticky="ew")

file_label = tk.StringVar(value="Файл для расшифровки: не выбран")
file_label_display = ttk.Label(frame_files, textvariable=file_label)
file_label_display.grid(row=0, column=0, sticky="w", padx=5, pady=5)

encrypt_file_label = tk.StringVar(value="Файл для шифрования: не выбран")
encrypt_file_label_display = ttk.Label(frame_files, textvariable=encrypt_file_label)
encrypt_file_label_display.grid(row=1, column=0, sticky="w", padx=5, pady=5)

frame_buttons_top = ttk.Frame(root, padding=10)
frame_buttons_top.grid(row=1, column=0, sticky="ew")

btn_select_file = ttk.Button(frame_buttons_top, text="Выбрать файл для расшифровки", command=chooseFileForDecryption)
btn_select_file.grid(row=0, column=0, padx=10, pady=5, sticky="ew")

btn_select_file_for_encryption = ttk.Button(frame_buttons_top, text="Выбрать файл для шифрования", command=chooseFileForEncryption)
btn_select_file_for_encryption.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

frame_buttons_main = ttk.Frame(root, padding=10)
frame_buttons_main.grid(row=2, column=0, sticky="ew")

btn_deobfuscate = ttk.Button(frame_buttons_main, text="Расшифровать", command=initiateDecryptionProcess, state=tk.DISABLED)
btn_deobfuscate.grid(row=0, column=0, padx=80, pady=10, sticky="ew")

md5_button = ttk.Button(frame_buttons_main, text="MD5", command=hash_MD5, width=10, state=tk.DISABLED)
md5_button.grid(row=0, column=3, padx=5, pady=5)

sha1_button = ttk.Button(frame_buttons_main, text="SHA-1", command=hash_SHA1, width=10, state=tk.DISABLED)
sha1_button.grid(row=0, column=4, padx=5, pady=5)

sha256_button = ttk.Button(frame_buttons_main, text="SHA-256", command=hash_SHA256, width=10, state=tk.DISABLED)
sha256_button.grid(row=1, column=3, padx=5, pady=5)

sha512_button = ttk.Button(frame_buttons_main, text="SHA-512", command=hash_SHA512, width=10, state=tk.DISABLED)
sha512_button.grid(row=1, column=4, padx=5, pady=5)

root.update_idletasks()
root.geometry('550x290')
setWindow(root)

root.mainloop()

